import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows


class DynamicFieldCSVReader:
    def __init__(self, csv_path):
        self.csv_path = csv_path
        self.df = pd.read_csv(csv_path)
        self.dynamic_fields = {
            "DAS": {"count": 0, "length": 8, "total_length": 0, "count_position": (175, 176)},
            "CSARR": {"count": 0, "length": None, "total_length": 0, "count_position": (177, 179)},
            "CCAM": {"count": 0, "length": None, "total_length": 0, "count_position": (180, 181)}
        }
        self.dynamic_start_position = 187
        self.format_info = {}
        self._prepare_format_info()

    def _prepare_format_info(self):
        # Initialize format_info with static fields and prepare dynamic fields structure
        for _, row in self.df.iterrows():
            var_name, start, end, length = row['Libellé des variables'], row.get('Position début'), row.get('Position fin'), row.get('Taille')

            # Skip if positions are not provided, indicating a dynamic field's sub-field
            if pd.isna(start) or pd.isna(end):
                self._add_dynamic_field(var_name, length)
                continue

            # Handle static fields
            try:
                start, end = int(start), int(end)
                self.format_info[var_name] = (start, end)
            except ValueError:
                continue

    def _add_dynamic_field(self, var_name, length):
        # Update dynamic fields structure based on var_name prefix
        for field_name in self.dynamic_fields:
            if var_name.startswith(field_name):
                if 'sub_fields' not in self.dynamic_fields[field_name]:
                    self.dynamic_fields[field_name]['sub_fields'] = []
                self.dynamic_fields[field_name]['sub_fields'].append((var_name, int(length)))
                # Dynamically update the total length of dynamic fields
                self.dynamic_fields[field_name]['length'] = sum(l for _, l in self.dynamic_fields[field_name]['sub_fields'])

    def _update_dynamic_counts(self, text_line):
        # Update counts based on positions defined in the class initialization
        for field in self.dynamic_fields.keys():
            p1, p2 = self.dynamic_fields[field]['count_position']
            count = int(text_line[p1:p2].strip())
            self.dynamic_fields[field]['count'] = count
            self.dynamic_fields[field]['total_length'] = count * self.dynamic_fields[field]['length']

    def generate_dynamic_fields_positions(self):
        sum_length = self.dynamic_start_position
        for field_key, field in self.dynamic_fields.items():
            count = field['count']
            for i in range(count):
                sub_length = sum_length
                for sub_field_name, length in field.get('sub_fields', []):
                    # Extract the descriptor part of the sub-field name (after the main field name)
                    descriptor = sub_field_name.split(field_key)[1]
                    # Format the name with the index after the field key and before the descriptor
                    var_name = f"{field_key}-{i+1} {descriptor}"
                    self.format_info[var_name] = (sub_length+1, sub_length + length)
                    sub_length += length
                sum_length += field['length']  # Move to the start position of the next set of dynamic fields

    @staticmethod
    def _convert_date(date_str):
        try:
            # Assuming date fields are in the format yyyymmdd
            return datetime.datetime.strptime(date_str, '%d%m%Y').strftime('%d/%m/%Y')
        except ValueError:
            # If conversion fails return the original string
            return date_str

    def extract_info(self, text_line):
        self._update_dynamic_counts(text_line)
        self.generate_dynamic_fields_positions()
        info = {}

        for var_name, (start, end) in self.format_info.items():
            value = text_line[start-1:end].strip() if start is not None and end is not None else "Not Found"
            # Check if this is a date field and convert it if it is
            if 'Date' in var_name:
                value = self._convert_date(value)
            info[var_name] = value

        return info

    def process_file(self, file_path):
        with open(file_path, 'r') as file:
            lines = file.readlines()

        # Create a list to store extracted information for each line
        extracted_data = []
        for line in lines:
            info = self.extract_info(line)
            extracted_data.append(info)

        # Convert the list of dictionaries into a DataFrame
        return pd.DataFrame(extracted_data)

    @staticmethod
    def apply_styles(workbook, df, exclusion_list=None):
        """
        Apply styles to the workbook based on the dataframe's data types and exclusion list.

        :param workbook: The openpyxl Workbook object
        :param df: DataFrame with the extracted information
        :param exclusion_list: List of columns to exclude from styling
        """
        if exclusion_list is None:
            exclusion_list = []

        # Access the first worksheet in the workbook
        worksheet = workbook.active

        # Define styles
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
        date_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")  # Light red
        int_fill = PatternFill(start_color="99CCFF", end_color="99CCFF", fill_type="solid")  # Light blue

        # Apply header styles
        for col_num, column_title in enumerate(df.columns, start=1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            if column_title in exclusion_list:
                worksheet.column_dimensions[cell.column_letter].hidden = True

        # Apply data styles
        for col_num, column_title in enumerate(df.columns, start=1):
            if column_title in exclusion_list:
                continue
            for row_num, item in enumerate(df[column_title], start=2):
                cell = worksheet.cell(row=row_num, column=col_num)
                if 'Date' in column_title:
                    cell.fill = date_fill
                elif isinstance(item, int):
                    cell.fill = int_fill

    def generate_and_style_excel(self, file_path, output_excel_path, exclusion_list=None):
        """
        Generate an Excel file from a text file and apply styles to it.

        :param file_path: Path to the text file to process
        :param output_excel_path: Path where the styled Excel file will be saved
        :param exclusion_list: List of column headers to exclude from the Excel file
        """
        # First, generate the DataFrame by processing the text file
        df = self.process_file(file_path)

        # Define colors for headers and data types
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
        date_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")  # Light red
        int_fill = PatternFill(start_color="99CCFF", end_color="99CCFF", fill_type="solid")  # Light blue

        # Now, use Pandas Excel writer with the openpyxl engine
        with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:
            # Write DataFrame to Excel
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            # Get the openpyxl workbook and worksheet objects
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']

            # Apply header styles
            for col in worksheet.iter_cols(min_row=1, max_row=1, max_col=len(df.columns)):
                for cell in col:
                    if cell.value not in exclusion_list:
                        cell.font = header_font
                        cell.fill = header_fill

            # Apply data styles
            for col in worksheet.iter_cols(min_row=2, max_col=len(df.columns)):
                for cell in col:
                    if cell.value is not None:
                        column_letter = cell.column_letter
                        header_value = worksheet[f"{column_letter}1"].value
                        if 'Date' in header_value:
                            cell.fill = date_fill
                        elif isinstance(cell.value, int):
                            cell.fill = int_fill
                        if header_value in exclusion_list:
                            worksheet.column_dimensions[column_letter].hidden = True


# Example usage:
csv_path = './assets/format_RHS-2023_jugu.csv'
format_reader = DynamicFieldCSVReader(csv_path)
rhs_file_path = './rhs/RHS_Groupe_20231019,1441.txt'
output_excel_path = './output/RHS_extracted_data_styled.xlsx'
exclusion_list = ['Filler1', 'Filler2']

# Generate and style the Excel file
format_reader.generate_and_style_excel(rhs_file_path, output_excel_path, exclusion_list)

print(f"Data extracted and written to styled Excel file at: {output_excel_path}")


# Rest of the class remains unchanged ...

