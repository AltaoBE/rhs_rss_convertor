import datetime
import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
from collections import OrderedDict

from openpyxl.utils import get_column_letter


class DynamicFieldCSVReader:
    def __init__(self, csv_path):
        self.csv_path = csv_path
        self.df = pd.read_csv(csv_path)
        self.dynamic_fields = {
            "DAS": {"count": 0, "length": 8, "total_length": 0, "count_position": (175, 176)},
            "CSARR": {"count": 0, "length": 35, "total_length": 0, "count_position": (177, 179)},
            "CCAM": {"count": 0, "length": 23, "total_length": 0, "count_position": (180, 181)}
        }
        self.dynamic_start_position = 187
        self.format_info = OrderedDict()
        self._prepare_format_info()

    def _prepare_format_info(self):
        # Initialize format_info with static fields and prepare dynamic fields structure
        for _, row in self.df.iterrows():
            var_name, start, end, length = row['Libellé des variables'], row.get('Position début'), row.get(
                'Position fin'), row.get('Taille')

            # Skip if positions are not provided, indicating a dynamic field's sub-field
            if pd.isna(start) or pd.isna(end):
                self._add_dynamic_field(var_name, length)
                continue

            # Handle static fields
            try:
                start, end = int(start), int(end)
                self.format_info[var_name] = (start, end)
            except ValueError:
                continue

    def _add_dynamic_field(self, var_name, length):
        # Update dynamic fields structure based on var_name prefix
        for field_name in self.dynamic_fields:
            if var_name.startswith(field_name):
                if 'sub_fields' not in self.dynamic_fields[field_name]:
                    self.dynamic_fields[field_name]['sub_fields'] = []
                self.dynamic_fields[field_name]['sub_fields'].append((var_name, int(length)))
                # Dynamically update the total length of dynamic fields
                self.dynamic_fields[field_name]['length'] = sum(
                    l for _, l in self.dynamic_fields[field_name]['sub_fields'])

    def _update_dynamic_counts(self, text_line):
        # Update counts based on positions defined in the class initialization
        for field in self.dynamic_fields.keys():
            p1, p2 = self.dynamic_fields[field]['count_position']
            count = int(text_line[p1 - 1:p2].strip())
            self.dynamic_fields[field]['count'] = count
            self.dynamic_fields[field]['total_length'] = count * self.dynamic_fields[field]['length']

    def generate_dynamic_fields_positions(self):
        sum_length = self.dynamic_start_position
        for field_key, field in self.dynamic_fields.items():
            count = field['count']
            for i in range(count):
                sub_length = sum_length
                for sub_field_name, length in field.get('sub_fields', []):
                    # Extract the descriptor part of the sub-field name (after the main field name)
                    descriptor = sub_field_name.split(field_key)[1]
                    # Format the name with the index after the field key and before the descriptor
                    var_name = f"{field_key}-{i + 1} {descriptor}"
                    self.format_info[var_name] = (sub_length, sub_length + length)
                    sub_length += length

                sum_length += field['length']  # Move to the start position of the next set of dynamic fields

    @staticmethod
    def _convert_date(date_str):
        try:
            # Assuming date fields are in the format yyyymmdd
            return datetime.datetime.strptime(date_str, '%d%m%Y').strftime('%d/%m/%Y')
        except ValueError:
            # If conversion fails return the original string
            return date_str

    def extract_info(self, text_line):
        self._update_dynamic_counts(text_line)
        self.generate_dynamic_fields_positions()
        info = {}

        for var_name, (start, end) in self.format_info.items():
            value = text_line[start-1:end] if start is not None and end is not None else "Not Found"
            # Check if this is a date field and convert it if it is
            if 'Date' in var_name:
                value = self._convert_date(value)
            info[var_name] = value

        return info

    def process_file(self, file_path):
        with open(file_path, 'r') as file:
            lines = file.readlines()

        # Lists to store extracted information for each line and the special fields
        extracted_data = []
        das_fields = []
        csarr_fields = []
        ccam_fields = []

        # Process each line to extract information
        for line in lines:
            info = self.extract_info(line)

            # Separate out the DAS, CSARR, and CCAM fields
            das_info = {k: v for k, v in info.items() if k.startswith('DAS')}
            csarr_info = {k: v for k, v in info.items() if k.startswith('CSARR')}
            ccam_info = {k: v for k, v in info.items() if k.startswith('CCAM')}

            # Remove these fields from the main info dictionary
            for k in list(das_info) + list(csarr_info) + list(ccam_info):
                del info[k]

            # Append the main info and special fields to their respective lists
            extracted_data.append(info)
            das_fields.append(das_info)
            csarr_fields.append(csarr_info)
            ccam_fields.append(ccam_info)

        # Convert the list of dictionaries into DataFrames
        extracted_df = pd.DataFrame(extracted_data)
        das_df = pd.DataFrame(das_fields)
        csarr_df = pd.DataFrame(csarr_fields)
        ccam_df = pd.DataFrame(ccam_fields)

        # Concatenate the DataFrames to get them in the desired order
        final_df = pd.concat([extracted_df, das_df, csarr_df, ccam_df], axis=1)

        return final_df

    def generate_excel(self, output_path, rhs_file_path, exclusion_list=None, column_width=None):  #
        if exclusion_list is None:
            exclusion_list = []

        # Remove excluded columns
        dataframe = self.process_file(rhs_file_path)
        dataframe = dataframe.drop(columns=exclusion_list, errors='ignore')

        # Create a Pandas Excel writer using openpyxl as the engine
        writer = pd.ExcelWriter(output_path, engine='openpyxl')
        dataframe.to_excel(writer, index=False, sheet_name='Extracted Data')

        # Get the openpyxl workbook and worksheet objects
        workbook = writer.book
        worksheet = writer.sheets['Extracted Data']

        # Set the column width and text wrap
        for col in worksheet.columns:
            for cell in col:
                cell.alignment = Alignment(wrap_text=True)
            if column_width is not None:
                worksheet.column_dimensions[get_column_letter(col[0].column)].width = column_width
            else:
                worksheet.column_dimensions[get_column_letter(col[0].column)].auto_size = True

        # Save the workbook
        writer.close()

# Example usage:


# Example usage:
# csv_path = './assets/format_RHS-2023_jugu.csv'
# format_reader = DynamicFieldCSVReader(csv_path)
# rhs_file_path = './input/RHS_Groupe_20231019,1441.txt'
# output_excel_path = './output/RHS_extracted_data.xlsx'
# exclusion_list = ['Filler1', 'Filler2']
#
# # Generate and style the Excel file
# format_reader.generate_excel(output_excel_path, exclusion_list)
#
# print(f"Data extracted and written to styled Excel file at: {output_excel_path}")
